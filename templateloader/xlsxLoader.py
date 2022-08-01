from openpyxl import load_workbook
from pandas import DataFrame
import xlsxwriter


# 엑셀 업로드
class xlsxUpload():
    # 엔티티
    def entity(cdVal):
        # 파일 불러오기
        wb = load_workbook('templateloader/dataset/' + cdVal + '.xlsx', data_only=True)

        # 시트 불러오기
        ws = wb['entity']
        
        # global
        dataAll = []

        # ws 이터레이터
        for row in ws.iter_rows():

            # 한 줄 담기
            dataRow = []
            for cell in row:
                dataRow.append(cell.value)

            # 전체 리스트(dataAll)에 담기
            dataAll.append(dataRow)
        return dataAll

    # 인텐트
    def intent(cdVal):
        # 파일 불러오기
        wb = load_workbook('templateloader/dataset/' + cdVal + '.xlsx', data_only=True)

        # 시트 불러오기
        ws = wb['intent']
        
        # global
        dataAll = []

        # ws 이터레이터
        for row in ws.iter_rows():

            # 한 줄 담기
            dataRow = []
            for cell in row:
                dataRow.append(cell.value)

            # 전체 리스트(dataAll)에 담기
            dataAll.append(dataRow)
        return dataAll


# 엑셀 다운로드
class xlsxDownload():
    # 엔티티
    def entity(cdVal, entity_list, synonym_list):

        # 헤더 생성
        header = ["엔티티명", "대표어", "동의어"]

        # 데이터 입력
        data_list = []
        tmpEntity = []
        for item in entity_list:
            tmpSyn = ""
            for synonym in synonym_list:
                if synonym.entity_id == item.id:
                    if tmpSyn == "":
                        tmpSyn = tmpSyn + synonym.entry
                    else:
                        tmpSyn = tmpSyn + "\n" + synonym.entry
            tmpEntity.append(item.subject)
            tmpEntity.append(item.entry)
            tmpEntity.append(tmpSyn)
            data_list.append(tmpEntity)
            tmpEntity = []
        
        # 저장
        dataF = DataFrame(data_list,columns=header)
        filename = 'D:/download/' + cdVal + '_entity.xlsx'
        dataF.to_excel(filename, 'Sheet',index=False, engine='xlsxwriter')

        return print('download: ' + cdVal + '_entity.xlsx')

    # 인텐트
    def intent(cdVal, intent_list, sentence_list, response_list):

        # 헤더 생성
        header = ["인텐트명", "등록예문", "응답"]

        # 데이터 입력
        data_list = []
        tmpIntent = []
        for item in intent_list:
            tmpSent = ""
            tmpRes = ""
            for sentence in sentence_list:
                if sentence.intent_id == item.id:
                    if tmpSent == "":
                        tmpSent = tmpSent + sentence.phrase
                    else:
                        tmpSent = tmpSent + "\n" + sentence.phrase
            for response in response_list:
                if response.intent_id == item.id:
                    tmpRes = response.phrase
            tmpIntent.append(item.subject)
            tmpIntent.append(tmpSent)
            tmpIntent.append(tmpRes)
            data_list.append(tmpIntent)
            tmpIntent = []
        
        # 저장
        dataF = DataFrame(data_list,columns=header)
        filename = 'D:/download/' + cdVal + '_intent.xlsx'
        dataF.to_excel(filename, 'Sheet',index=False, engine='xlsxwriter')

        return print('download: ' + cdVal + '_intent.xlsx')